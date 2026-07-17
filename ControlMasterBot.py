import telebot 
from telebot import types 
import threading
import tempfile
from PIL import ImageGrab
import os 
import pyautogui as pg
import keyboard
import webbrowser
from pycaw.pycaw import AudioUtilities, IAudioEndpointVolume
from ctypes import cast, POINTER
from comtypes import CLSCTX_ALL, CoInitialize
import cv2
import platform
import psutil
import time
import logging
import configparser
import pyperclip
import mss
import numpy as np
import subprocess
import openpyxl

config = configparser.ConfigParser()
config.read('config.ini')

bot_token = config['bot']['token']

bot = telebot.TeleBot(bot_token, parse_mode="HTML")

pg.FAILSAFE = False 
recording = False
recording_thread = False
out = None
mouse_blocked = False
keyboard_blocked = False

admin_password = config['bot']['admin_password']
user_attempts = {}

def startup_logging():

    logging.basicConfig(
        filename = 'RemoteControlBot.log',
        filemode = 'a',
        level = logging.DEBUG,
        format = '%(asctime)s - %(levelname)s - %(message)s',
        encoding = 'utf-8'
    )

def send_message_to_chat(chat_id, text):
    bot.send_message(chat_id, text)

def send_message_async(chat_id, message_text):
    threading.Thread(target=send_message_to_chat, args=(chat_id, message_text)).start()

file_name = "black_list.xlsx"
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Chat-id', 'First Name'])
    wb.save(file_name)

def block_user_from_chat_id(chat_id: int):
    try:
        logging.info("blocking user has been started")

        chat = bot.get_chat(chat_id)
        first_name = getattr(chat, "first_name", None)
    
        file_name = "black_list.xlsx"
        if not os.path.exists(file_name):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['Chat - id', 'First Name'])
            wb.save(file_name)

        def user_exists(chat_id):
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == chat_id:
                    return True
                
            return False
        
        def save_user(chat_id, first_name):
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            sheet.append([chat_id, first_name])
            wb.save(file_name)

        if not user_exists(chat_id):
            save_user(chat_id, first_name)
            logging.info(f"new user is added {chat_id}")
        else:
            logging.info(f"User {chat_id} already registred, not added again")    

    except Exception as e:
        pass

def get_user_info_by_chat_id_and_add_to_base(chat_id: int):
    try:
        logging.info("get user data has been started")

        chat = bot.get_chat(chat_id)
        
        first_name = getattr(chat, "first_name", None)
        last_name = getattr(chat, "last_name", None)
        username = getattr(chat, "username", None)

        file_name = "data.xlsx"
        if not os.path.exists(file_name):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['Chat-id', 'First Name', 'Last Name', 'Username'])
            wb.save(file_name)

        def user_exists(chat_id):
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == chat_id:
                    return True
            return False

        def save_user(chat_id, first_name, last_name, username):
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
            sheet.append([chat_id, first_name, last_name, username])
            wb.save(file_name)

        if not user_exists(chat_id):
            save_user(chat_id, first_name, last_name, username)
            logging.info(f"New user is added {chat_id}")
        else:
            logging.info(f"User {chat_id} already registered, not added again")

    except Exception as e:
        logging.error(f"get user data error: {e}")

@bot.message_handler(commands=['start'])
def send_welcome(message):
    try:
        startup_logging()
        logging.info('bot has been started')

        chat_id = message.chat.id

        wb = openpyxl.load_workbook("black_list.xlsx")
        sheet = wb.active

        found = False
        for cell in sheet["A"]:
            if str(chat_id) == str(cell.value):
                found = True
                break

        if found:
            BreakSystem()

        print(f"Chat ID: {chat_id}")
        bot.send_message(chat_id, '👋 Welcome!')

        get_user_info_by_chat_id_and_add_to_base(chat_id)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("Keyboard")
        btn2 = types.KeyboardButton("Mouse")
        btn3 = types.KeyboardButton("System")
        btn4 = types.KeyboardButton("screenshot")
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, '👋', reply_markup=markup)

    except Exception as e:
        logging.error('Starting error')

def BreakSystem():
    try:
        logging.info('break system has been started')

        os.system('taskkill /f /im python.exe')

    except Exception as e:
        logging.error('bot breaking error')

def handle_screenshot(message):
    try:
        logging.info('screenshot command has been started')

        path = os.path.join(tempfile.gettempdir(), 'screenshot.png')
        screenshot = ImageGrab.grab()
        screenshot.save(path, 'PNG')

        with open(path, 'rb') as photo:
            bot.send_photo(message.chat.id, photo)

        logging.info('Screenshot has been sended')

    except Exception as e:
        logging.error(f'screen shot error {e}')
        bot.send_message(message.chat.id, "Screenshot error")

def handle_shutdown(message):
    try:
        logging.info('system shutdown has been started')

        bot.send_message(message.chat.id, 'shutdown...')

        with open("RemoteControlBot.log", "rb") as log_file:
                bot.send_document(message.chat.id, log_file)

        os.system("shutdown -s -t 1")
        bot.send_message(message.chat.id, "system is off")

    except Exception as e:
        logging.error('System shutdown error')

def handle_reboot(message):
    try:
        logging.info('system reboot has been started')

        bot.send_message(message.chat.id, "reboot system ...")
        os.system("shutdown -r -t 1")
        bot.send_message(message.chat.id, "the system has started rebooting")
    
    except Exception as e:
        logging.error('system reboot error')

def clipboard(message):
    try:
        logging.info('clipboard command has been started')

        x = pyperclip.paste()
        bot.send_message(message.chat.id, x)

        logging.info('command is complete')

    except Exception as e:
        logging.error('Clipboard command error')

def OpenLink(message):
    try:
        logging.info('open link command has been started')

        bot.send_message(message.chat.id, "please send me the link")
        def save_link(msg):
            try:
                link = msg.text
                logging.info('link has been saved')

            except Exception as e:
                logging.error('link save error')

            bot.send_message(msg.chat.id, f"Opening link: {link}")
            webbrowser.open_new(link)
        bot.register_next_step_handler(message, save_link)

    except Exception as e:
        logging.error('open link command error')

def screen_record():
    global recording, out

    with mss.mss() as sct:
        monitor = sct.monitors[1]
        width = monitor["width"]
        height = monitor["height"]

        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        out = cv2.VideoWriter("video.mp4", fourcc, 10.0, (width, height))

        while recording:
            img = sct.grab(monitor)
            frame = np.array(img)
            frame = cv2.cvtColor(frame, cv2.COLOR_BGRA2BGR)
            out.write(frame)

        out.release()
        out = None

def start_recording(message):
    global recording, recording_thread

    try:
        logging.info('Recording has been started')

        if not recording:
            recording = True

            recording_thread = threading.Thread(target=screen_record)
            recording_thread.start()

            logging.info('Rec.')
            bot.reply_to(message, 'Recording has been started')
        
        else:
            logging.info('Rec.')

            bot.reply_to(message, 'The recording is already connected.')

    except Exception as e:
        logging.error('Rec. error')

def stop_recording(message):
    global recording

    try:
        logging.info('Rec. break command has been started')

        if recording:
            recording = False

            logging.info('Rec. breaked , file has been sended')
            bot.reply_to(message, 'Recording has been stopeed, file saved')

        else:
            logging.info('Rec. breaked')
            bot.reply_to(message, 'The recording is already stopped')

    except Exception as e:
        logging.error('Breaking rec error')

def keyboard_button(message):
    try:
        logging.info('any button command has been sendecd')
        
        bot.send_message(message.chat.id, "please say me button name")
        def save_button_name(msg):
            try:
                button = msg.text
                logging.info('Button has been saved')
            
            except Exception as e:
                logging.error('button save error')
            
            pg.press(button)
            bot.send_message(message.chat.id, f"Button {button} pressed")
        bot.register_next_step_handler(message, save_button_name)

    except Exception as e:
        logging.error('any button command error')

def Vc(message):
    try:
        logging.info('voice control command is sended')

        bot.send_message(message.chat.id, "please say the volume procentage")
        def save_value(msg):
            
            volume_procent = float(msg.text) / 100
            logging.info('voice value has been saved')

            CoInitialize()
            devices = AudioUtilities.GetSpeakers()
            interface = devices.Activate(IAudioEndpointVolume._iid_, CLSCTX_ALL, None)
            volume = cast(interface, POINTER(IAudioEndpointVolume))
            volume.SetMasterVolumeLevelScalar(volume_procent, None)
            bot.send_message(msg.chat.id, f"volume set to {msg.text}%")
            
        bot.register_next_step_handler(message, save_value)
    
    except Exception as e:
        logging.error('voice control command error')

def send_camera_photo(message):
    try:
        logging.info('camera photo command has been started')

        bot.send_message(message.chat.id, 'your camera photo')
        cam = cv2.VideoCapture(0)
        ret, frame = cam.read()
        if ret:
            photo_path = 'camera.jpg'
            cv2.imwrite(photo_path, frame)
            cam.release()

            with open(photo_path, 'rb') as photo:
                bot.send_photo(message.chat.id, photo)

            logging.info('Photo has been sended')

        else:
            logging.error('photo error')
    
    except Exception as e:
        logging.error('camera photo command error')
    
def block_mouse():
    try:
        logging.info('block mouse command has been started')

        x, y = pg.position()  

        while mouse_blocked:
            pg.moveTo(x, y)
            time.sleep(0.01)

    except Exception as e:
        logging.error(f'block mouse error: {e}')


def system_info(message):
    try:
        logging.info('system info command has been started')

        try:
            system = platform.system()
            node = platform.node()
            release = platform.release()
            version = platform.version()
            architecture = platform.architecture()

            cpu_usage = psutil.cpu_percent(interval=1)
            cpu_cores = psutil.cpu_count(logical=False)

            mem = psutil.virtual_memory()
            ram_usage_percent = mem.percent
            total_ram_gb = mem.total / (1024 ** 3)
            available_ram_gb = mem.available / (1024 ** 3)

            disk = psutil.disk_usage('C:\\')
            total_disk_gb = disk.total / (1024 ** 3)
            used_disk_gb = disk.used / (1024 ** 3)
            free_disk_gb = disk.free / (1024 ** 3)
            disk_percent = disk.percent

            logging.info("Data's has been saved")

        except Exception as e:
            logging.error("Data's save error")

        bot.send_message(message.chat.id, "this is your system info")
        bot.send_message(message.chat.id, "👇")

        bot.send_message(message.chat.id,
            f"System: {system}\n"
            f"Node Name: {node}\n"
            f"OS Release: {release}\n"
            f"Version: {version}\n"
            f"Architecture: {architecture}"
        )

        bot.send_message(message.chat.id, 
            f"CPU Usage: {cpu_usage}%\n"
            f"Number of Cores: {cpu_cores}\n"
        )

        bot.send_message(message.chat.id, 
            f"RAM Usage: {ram_usage_percent}%\n"
            f"Total RAM: {total_ram_gb:.2f} GB\n"
            f"Available RAM: {available_ram_gb:.2f} GB\n"
        )

        bot.send_message(message.chat.id, 
            f"Disk Size: {total_disk_gb:.2f} GB\n"
            f"Used: {used_disk_gb:.2f} GB\n"
            f"Free: {free_disk_gb:.2f} GB\n"
            f"Usage Percentage: {disk_percent}%"
        )

    except Exception as e:
        logging.error('system info command error')

def wifi_off(message):

    try:
        logging.info('wifi off command has been started')

        subprocess.call('netsh interface set interface name=\"Wi-Fi\" admin=disabled', shell=True)
        bot.send_message(message.chat.id, 'wifi is off')
        logging.info('wifi is off')

    except Exception as e:
        logging.error('wifi off command error')

def wifi_on(message):

    try:
        logging.info('wifi on command has been started')

        subprocess.call('netsh interface set interface name=\"Wi-Fi\" admin=enabled', shell=True)
        bot.send_message(message.chat.id, "wifi is on")
        logging.info('wifi is on ')

    except Exception as e:
        logging.error('wifi on command errorլ')

def block_keyboard():
    try:
        logging.info('block_keyboard command has been started')

        global keyboard_blocked  
        blocked_keys = [
            "q", "w", "e", "r", "t", "y", "u", "i", "o", "p", "[", "]",
            "a", "s", "d", "f", "g", "h", "j", "k", "l", ";", "'", "\n",
            "z", "x", "c", "v", "b", "n", "m", ",", ".", "/",
            "0", "1", "2", "3", "4", "5", "6", "7", "8", "9",
            "*", "-", "=", "+", "|", "`",
            "\t", " ", "shift", "windows", "alt", "esc", "backspace", "ctrl"
        ]  
        if keyboard_blocked:
            for key in blocked_keys:
                keyboard.block_key(key)
        else:
            for key in blocked_keys:
                keyboard.unblock_key(key)
    
    except Exception as e:
        logging.error('block keyboard error')

def handle_mouse_off(message):
    try:
        logging.info('mouse block command has been started')
        global mouse_blocked
        if not mouse_blocked:
            mouse_blocked = True
            bot.send_message(message.chat.id, 'mouse has been blocked')
            threading.Thread(target=block_mouse, daemon=True).start()

    except Exception as e:
        logging.error('mouse block command error')
    
def handle_mouse_on(message):
    try:
        logging.info('mouse on command has been started')
        
        global mouse_blocked
        mouse_blocked = False
        bot.send_message(message.chat.id, 'mouse has been unblocked')
        logging.info('mouse has been unblocked')

    except Exception as e:
        logging.error('mouse on command error')

def handler_keyboard_off(message):
    try:
        logging.info('keyboard off command has been started')
        global keyboard_blocked
        keyboard_blocked = True
        bot.send_message(message.chat.id, 'keyboard has been blocked')
        block_keyboard()
        logging.info('keyboard has been blocked')

    except Exception as e:
        logging.error('keyboard off command error')

def handler_keyboard_on(message):
    try:
        logging.info('keyboard on command has been started')

        global keyboard_blocked
        keyboard_blocked = False
        bot.send_message(message.chat.id, 'keyboard has been unblocked')
        block_keyboard()

        logging.info('keyboard has been unblocked')

    except Exception as e:
        logging.error('keyboard on command error')

def handler_OpenLink(message):
    try:
        logging.info('open link has been started')

        OpenLink()
        bot.send_message(message.chat.id, 'link has been opened')

        logging.info('link has been opened')
    
    except Exception as e:
        logging.error('open link command error')

@bot.message_handler(commands=['log'])
def send_logs(message):
    chat_id = message.chat.id
    user_attempts[chat_id] = 3  
    bot.send_message(chat_id, "enter the administrator password, you have 3 attempts")

    def save_pass(msg):
        attempt_password = msg.text

        if attempt_password == admin_password:
            bot.send_message(chat_id, "you typed the correct password")

            bot.send_message(chat_id, "sending bot log file")
            logging.info("log file send command has been started")

            with open("RemoteControlBot.log", "rb") as log_file:
                bot.send_message(chat_id, "this is your log file ⬇️")
                bot.send_document(chat_id, log_file)

            user_attempts.pop(chat_id, None)

        else:
            user_attempts[chat_id] -= 1
            attempts_left = user_attempts[chat_id]

            if attempts_left > 0:
                bot.send_message(chat_id, f"incorrect password ({attempts_left} attempts left)")
                bot.register_next_step_handler(msg, save_pass)
            else:
                bot.send_message(chat_id, "you don't have attempts, blocking the bot")
                BreakSystem()
                user_attempts.pop(chat_id, None)

    bot.register_next_step_handler(message, save_pass)

@bot.message_handler(commands={'del_log'})
def del_log(message):
    chat_id = message.chat.id
    user_attempts[chat_id] = 3
    bot.send_message(message.chat.id, 'enter the administrator password, you have 3 attempts')

    def save_pass(msg):
        attempt_password = msg.text

        if attempt_password == admin_password:

            bot.send_message(message.chat.id, 'deleting log file from pc')
            os.remove("ControlMasterBot.log")
            bot.send_message(message.chat.id, 'log file has been deleted')

        else:
            user_attempts[chat_id] -= 1
            attempts_left = user_attempts[chat_id]

            if attempts_left > 0:
                bot.send_message(chat_id, f"incorrect password ({attempts_left} attempts left)")
                bot.register_next_step_handler(msg, save_pass)

            else:
                bot.send_message(message.chat.id, "you don't have attempts, blocking bot")
                user_attempts.pop(chat_id, None)
                BreakSystem()
            
    bot.register_next_step_handler(message, save_pass)
    
@bot.message_handler(commands={'send_data_file'})
def send_base(message):
    chat_id = message.chat.id
    user_attempts[chat_id] = 3
    bot.send_message(message.chat.id, 'enter the administrator password, you have 3 attempts')

    def save_pass(msg):
        attempt_password = msg.text

        if attempt_password == admin_password:

            bot.send_message(message.chat.id, 'sending data file')
            logging.info("dat file sending has been started")

            with open("data.xlsx", "rb") as data_file:
                bot.send_message(message.chat.id, "this is your data file ⬇️")
                bot.send_document(message.chat.id, data_file)

            logging.info("data file has been sended")

        else:
            user_attempts[chat_id] -= 1
            attempts_left = user_attempts[chat_id]

            if attempts_left > 0:
                bot.send_message(chat_id, f"incorrect password ({attempts_left} attempts left)")
                bot.register_next_step_handler(msg, save_pass)

            else:
                bot.send_message(message.chat.id, "you don't have attempts, blocking bot")
                user_attempts.pop(chat_id, None)
                BreakSystem()
        
    bot.register_next_step_handler(message, save_pass)

@bot.message_handler(commands={"del_data_file"})
def del_base(message):
    chat_id = message.chat.id
    user_attempts[chat_id] = 3
    bot.send_message(message.chat.id, 'enter the administrator password, you have 3 attempts')

    def save_pass(msg):
        attempt_password = msg.text

        if attempt_password == admin_password:

            bot.send_message(message.chat.id, 'deleting data file')
            os.remove("data.xlsx")
            bot.send_message(message.chat.id, 'data file has been deleted')
        
        else:
            user_attempts[chat_id] -= 1
            attempts_left = user_attempts[chat_id]

            if attempts_left > 0:
                bot.send_message(chat_id, f"incorrect password ({attempts_left} attempts left)")
                bot.register_next_step_handler(msg, save_pass)

            else:
                bot.send_message(message.chat.id, "you don't have attempts, blocking bot")
                user_attempts.pop(chat_id, None)
                BreakSystem()
        
        bot.register_next_step_handler(message, save_pass)

@bot.message_handler(commands=['block_user'])
def blocking_user(message):
    chat_id = message.chat.id
    user_attempts[chat_id] = 3
    bot.send_message(chat_id, 'enter the administrator password, you have 3 attempts')

    def save_pass(msg):
        attempts_password = msg.text

        if attempts_password == admin_password:
            bot.send_message(chat_id, 'Please send me user chat_id')

            def save_id(msg2):
                user_id = msg2.text
                wb = openpyxl.load_workbook("data.xlsx")
                sheet = wb.active

                found = False
                for cell in sheet["A"]:
                    if str(user_id) == str(cell.value):
                        found = True
                        break

                if found:
                    block_user_from_chat_id(user_id)
                    bot.send_message(chat_id, f"User {user_id} has been blocked")
                else:
                    bot.send_message(chat_id, f"User {user_id} not found")

            bot.register_next_step_handler(msg, save_id) 

        else:
            user_attempts[chat_id] -= 1
            if user_attempts[chat_id] > 0:
                bot.send_message(chat_id, f"Wrong password, {user_attempts[chat_id]} attempts left")
                bot.register_next_step_handler(msg, save_pass)
            else:
                bot.send_message(chat_id, "No attempts left, access denied")

    bot.register_next_step_handler(message, save_pass)

@bot.message_handler(content_types=['text'])
def func(message):
    if message.text == "Keyboard":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("⌨ off")
        btn2 = types.KeyboardButton("⌨ on")
        btn3 = types.KeyboardButton("(Ctrl + W) close active window")
        btn4 = types.KeyboardButton("( space )")
        btn5 = types.KeyboardButton("(Win + D) to desktop")
        btn6 = types.KeyboardButton("home page")
        btn7 = types.KeyboardButton("screenshot")
        btn8 = types.KeyboardButton("any button")
        markup.add(btn1, btn2, btn3, btn3, btn4, btn5, btn6, btn7, btn8)
        bot.send_message(message.chat.id, 'keyboard control', reply_markup=markup)

    elif message.text == '(Ctrl + W) close active window':
        try:
            logging.info('(ctrl + w ) hotkey error')

            pg.hotkey('ctrl', 'w')
            bot.send_message(message.chat.id, '(Ctrl + W) hotkey has been pressed')
            logging.info('hotkey has been pressed')

        except Exception as e:
            logging.error('hotkey press error')

    elif message.text == '( space )':
        try:
            logging.info('space press command has bee started')

            pg.press('space')
            bot.send_message(message.chat.id, 'button (space) has been pressed')
            logging.info('space has been pressed')

        except Exception as e:
            logging.error('space press error')

    elif message.text == '(Win + D) to desktop':
        try:
            logging.info('win + d hotkey has been started')

            pg.hotkey('win', 'd')
            bot.send_message(message.chat.id, '(Win + D) hotkey has been pressed')
            logging.info('win + d has been pressed')

        except Exception as e:
            logging.error('win + d hotkey press error')

    elif message.text == 'any button':
        try:
            logging.info('any button command has been started')
            
            keyboard_button(message)
            logging.info('any button is complete')

        except Exception as e:
            logging.error('any button command error')

    elif message.text == "Mouse":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton('🖱 off')
        btn2 = types.KeyboardButton('🖱 on')
        btn3 = types.KeyboardButton('left click')
        btn4 = types.KeyboardButton('right click')
        btn5 = types.KeyboardButton('home page')
        btn6 = types.KeyboardButton('screenshot')
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
        bot.send_message(message.chat.id, 'mouse control', reply_markup=markup)

    elif message.text == 'left click':
        try:
            logging.info('left click command has been started')

            pg.click(button='left')
            logging.info('left click has been complete')
            bot.send_message(message.chat.id, 'left click has been clicked')

        except Exception as e:
            logging.error('left click error')

    elif message.text == 'right click':
        try:
            logging.info('right click command has been started')

            pg.click(button='right')
            logging.info('right click command has been complete')
            bot.send_message(message.chat.id, 'right click has been clicked')

        except Exception as e:
            logging.error('right click command error')

    elif message.text == 'System':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton('screenshot')
        btn2 = types.KeyboardButton('system off')
        btn3 = types.KeyboardButton('system reboot')
        btn4 = types.KeyboardButton('home page')
        btn5 = types.KeyboardButton('open link')
        btn6 = types.KeyboardButton('volume control')
        btn7 = types.KeyboardButton('system information')
        btn8 = types.KeyboardButton('camera photo')
        btn9 = types.KeyboardButton('clipboard')
        btn10 = types.KeyboardButton('start recording')
        btn11 = types.KeyboardButton('stop recording')
        markup.add(    
            btn1, btn2, btn3, btn4, btn5, btn6, btn7, btn8, btn9, btn10,
            btn11
         )
        bot.send_message(message.chat.id, 'system control', reply_markup=markup)

    elif message.text == 'home page':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton('Keyboard')
        btn2 = types.KeyboardButton('Mouse')
        btn3 = types.KeyboardButton("System")
        btn4 = types.KeyboardButton('screenshot')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'home page', reply_markup=markup)

    elif message.text == 'wifi on':
        wifi_on(message)

    elif message.text == 'wifi off':
        wifi_off(message)

    elif message.text == 'start recording':
        start_recording(message)

    elif message.text == 'stop recording':
        stop_recording(message)
        logging.info('Recording has been breaked, sending file ...')

        with open('video.mp4', 'rb') as video:
            bot.send_video(message.chat.id, video)
            
    elif message.text == 'camera photo':
        send_camera_photo(message)

    elif message.text == 'system information':
        system_info(message)

    elif message.text == 'clipboard':
        clipboard(message)
    
    elif message.text == 'open link':
        OpenLink(message)

    elif message.text == 'volume control':
        Vc(message)

    elif message.text == 'screenshot':
        handle_screenshot(message)

    elif message.text == 'system off':
        handle_shutdown(message)

    elif message.text == 'system reboot':
        handle_reboot(message)

    elif message.text == '🖱 off':
        handle_mouse_off(message)

    elif message.text == '🖱 on':
        handle_mouse_on(message)

    elif message.text == '⌨ off':
        handler_keyboard_off(message)

    elif message.text == '⌨ on':
        handler_keyboard_on(message)

    elif message.text == 'break':
        try:
            logging.info('break bot command has been started')
            
            bot.send_message(message.chat.id, 'bot has been breaked')
            logging.info('block bot command complete, log file sending')

            with open("RemoteControl.log", "rb") as log_file:
                bot.send_document(message.chat.id, log_file)

            BreakSystem()

        except Exception as e:
            logging.error('block bot command error')

bot.polling(non_stop=True)
